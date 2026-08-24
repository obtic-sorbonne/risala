"""
3_evaluation.py
Évaluation CER/WER + fiabilité des métadonnées pour le corpus Risala.

Sources :
  - Feuille "Manuel"  : une ligne par page (L1_CONT_1, L1_ENV_1...)
    → regroupées par lettre avant comparaison
  - Feuille "LLM"     : une ligne par lettre (sortie du script 1)

Modes disponibles (variable MODE) :
  "metadata"      → fiabilité des métadonnées uniquement
  "transcription" → CER/WER sur la transcription uniquement
  "all"           → les deux

Convention :
  - Une cellule vide dans Manuel = "pas encore relu" → ignorée
  - Seules les cellules renseignées dans Manuel entrent dans le calcul

Installation :
  pip install jiwer openpyxl --break-system-packages

Usage :
  python 3_evaluation.py
"""

import re
import sys
import json
import openpyxl
import jiwer
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from datetime import datetime
from collections import defaultdict
from pathlib import Path

# ==============================================================================
# CONFIG
# ==============================================================================
INPUT_FILE = "lettres.xlsx"
SHEET_MANUEL = "Manuel"
SHEET_LLM    = "LLM"

# Mode : "metadata" | "transcription" | "all"
MODE = "transcription"

# Dossier de sortie pour les fichiers générés
OUTPUT_DIR = Path(r"C:\Users\sorbonne\Documents\Workspace\Risala\output")

# Champs de métadonnées à évaluer
METADATA_FIELDS = [
    "sender_name",
    "sender_place",
    "date_sent",
    "recipient_name",
    "recipient_place",
    "date_received",
    "inmate_id",
]

# ==============================================================================
# TRANSFORMATION TEXTE pour CER/WER
# ==============================================================================
TEXT_TRANSFORM = jiwer.Compose([
    jiwer.RemoveMultipleSpaces(),
    jiwer.Strip(),
    jiwer.ReduceToListOfListOfWords(),
])

# ==============================================================================
# FONCTIONS COMMUNES
# ==============================================================================
def clean(val):
    """Nettoie une valeur : None si vide."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("nul", "none", "null", "", "#n/a", "nan", "nat"):
        return None
    return s

def fmt_date(val):
    """Convertit une valeur Excel en string YYYY-MM-DD ou None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d") if val.year >= 1940 else None
    s = str(val).strip()
    if s.lower() in ("nul", "none", "null", "", "#n/a", "nat"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d") if dt.year >= 1940 else None
        except ValueError:
            pass
    return s

def normalize(value):
    """Normalise pour comparaison : minuscules, espaces réduits."""
    return " ".join((value or "").strip().lower().split())

def clean_text(text):
    """Nettoie le texte pour CER/WER."""
    return (text or "").replace("\n", " ").replace("\r", " ").replace("\t", " ").strip()

def extract_prefix(letter_id):
    """L1_CONT_1 → L1"""
    if not letter_id:
        return None
    match = re.match(r'^(L\d+)', str(letter_id).strip(), re.IGNORECASE)
    return match.group(1) if match else str(letter_id).strip()

def is_cont(letter_id):
    return bool(re.search(r'CONT', str(letter_id or ""), re.IGNORECASE))

# ==============================================================================
# CHARGEMENT ET REGROUPEMENT DE LA FEUILLE MANUEL
# ==============================================================================
def load_manuel(wb):
    """
    Charge la feuille Manuel (une ligne par page) et regroupe par lettre.
    Retourne un dict : { (folder_id, letter_prefix) : { champ: valeur } }
    """
    ws = wb[SHEET_MANUEL]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    # Index des colonnes
    def col(name):
        return headers.index(name) if name in headers else None

    idx = {
        "folder_id":      col("folder_id"),
        "letter_id":      col("letter_id"),
        "sender_name":    col("sender_name"),
        "sender_place":   col("sender_place"),
        "date_sent":      col("date_sent"),
        "recipient_name": col("recipient_name"),
        "recipient_place":col("recipient_place"),
        "date_received":  col("date_received"),
        "inmate_id":      col("inmate_id"),
        "transcription":  col("transcription") or col("content"),  # compat ancien nom
        "notes":          col("notes"),
    }

    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        folder_id = clean(row[idx["folder_id"]]) if idx["folder_id"] is not None else None
        letter_id = clean(row[idx["letter_id"]]) if idx["letter_id"] is not None else None
        if not folder_id or not letter_id:
            continue
        prefix = extract_prefix(letter_id)
        groups[(folder_id, prefix)].append({
            "letter_id":      letter_id,
            "sender_name":    clean(row[idx["sender_name"]]),
            "sender_place":   clean(row[idx["sender_place"]]),
            "date_sent":      fmt_date(row[idx["date_sent"]]),
            "recipient_name": clean(row[idx["recipient_name"]]),
            "recipient_place":clean(row[idx["recipient_place"]]),
            "date_received":  fmt_date(row[idx["date_received"]]),
            "inmate_id":      clean(row[idx["inmate_id"]]),
            "transcription":  clean(row[idx["transcription"]]),
            "is_cont":        is_cont(letter_id),
        })

    # Regroupement : métadonnées = première valeur non vide, transcription = concat CONT
    result = {}
    for (folder_id, prefix), pages in groups.items():
        def first(field):
            for p in pages:
                if p[field]:
                    return p[field]
            return None
        cont_pages = [p for p in pages if p["is_cont"] and p["transcription"]]
        transcription = " ".join(p["transcription"] for p in cont_pages)
        result[(folder_id, prefix)] = {
            "sender_name":    first("sender_name"),
            "sender_place":   first("sender_place"),
            "date_sent":      first("date_sent"),
            "recipient_name": first("recipient_name"),
            "recipient_place":first("recipient_place"),
            "date_received":  first("date_received"),
            "inmate_id":      first("inmate_id"),
            "transcription":  transcription.strip() or None,
        }
    return result

# ==============================================================================
# CHARGEMENT DE LA FEUILLE LLM
# ==============================================================================
def load_llm(wb):
    """
    Charge la feuille LLM (une ligne par lettre).
    Retourne un dict : { (folder_id, letter_id) : { champ: valeur } }
    """
    ws = wb[SHEET_LLM]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    def col(name):
        return headers.index(name) if name in headers else None

    idx = {
        "folder_id":      col("folder_id"),
        "letter_id":      col("letter_id"),
        "sender_name":    col("sender_name"),
        "sender_place":   col("sender_place"),
        "date_sent":      col("date_sent"),
        "recipient_name": col("recipient_name"),
        "recipient_place":col("recipient_place"),
        "date_received":  col("date_received"),
        "inmate_id":      col("inmate_id"),
        "transcription":  col("transcription") or col("content"),
    }

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        folder_id = clean(row[idx["folder_id"]]) if idx["folder_id"] is not None else None
        letter_id = clean(row[idx["letter_id"]]) if idx["letter_id"] is not None else None
        if not folder_id or not letter_id:
            continue
        result[(folder_id, letter_id)] = {
            "sender_name":    clean(row[idx["sender_name"]]),
            "sender_place":   clean(row[idx["sender_place"]]),
            "date_sent":      fmt_date(row[idx["date_sent"]]),
            "recipient_name": clean(row[idx["recipient_name"]]),
            "recipient_place":clean(row[idx["recipient_place"]]),
            "date_received":  fmt_date(row[idx["date_received"]]),
            "inmate_id":      clean(row[idx["inmate_id"]]),
            "transcription":  clean(row[idx["transcription"]]),
        }
    return result

# ==============================================================================
# ÉVALUATION MÉTADONNÉES
# ==============================================================================
def evaluate_metadata(pairs):
    print("\n=== Fiabilité des métadonnées ===")
    for field in METADATA_FIELDS:
        total, correct, skipped = 0, 0, 0
        for lid, ref_row, hyp_row in pairs:
            ref = ref_row.get(field)
            hyp = hyp_row.get(field)
            if ref:
                total += 1
                if normalize(ref) == normalize(hyp):
                    correct += 1
            elif hyp:
                skipped += 1
        if total == 0:
            print(f"  {field:20s}: aucune donnée dans Manuel")
            continue
        pct = correct / total * 100
        extra = f"  ({skipped} non relue(s))" if skipped else ""
        print(f"  {field:20s}: {correct:3d}/{total:3d} corrects ({pct:5.1f}%){extra}")

# ==============================================================================
# ÉVALUATION TRANSCRIPTION (CER/WER)
# ==============================================================================
def evaluate_transcription(pairs):
    print("\n=== Transcription (CER / WER) ===")
    ids, refs, hyps = [], [], []
    for lid, ref_row, hyp_row in pairs:
        ref = clean_text(ref_row.get("transcription"))
        hyp = clean_text(hyp_row.get("transcription"))
        if not ref:
            continue
        ids.append(lid)
        refs.append(ref)
        hyps.append(hyp)

    if not refs:
        print("  Aucune transcription de référence dans Manuel.")
        print("  → Renseignez la colonne 'transcription' dans les pages CONT.")
        return

    cer_global = jiwer.cer(refs, hyps,
                           reference_transform=TEXT_TRANSFORM,
                           hypothesis_transform=TEXT_TRANSFORM)
    wer_global = jiwer.wer(refs, hyps,
                           reference_transform=TEXT_TRANSFORM,
                           hypothesis_transform=TEXT_TRANSFORM)

    print(f"  Évalué sur {len(refs)} lettre(s)")
    print(f"  CER global = {cer_global * 100:.1f}%")
    print(f"  WER global = {wer_global * 100:.1f}%")
    print("  (ponctuation et diacritiques comptent comme des erreurs)")

    print("\n  Détail par lettre (CER, du plus élevé au plus bas) :")
    per_letter = sorted(
        [(lid, jiwer.cer([r], [h],
                         reference_transform=TEXT_TRANSFORM,
                         hypothesis_transform=TEXT_TRANSFORM))
         for lid, r, h in zip(ids, refs, hyps)],
        key=lambda x: x[1], reverse=True
    )
    for lid, c in per_letter:
        print(f"    {lid}: CER = {c * 100:.1f}%")

# ==============================================================================
# MAIN
# ==============================================================================
# ==============================================================================
# EXPORT RÉSULTATS
# ==============================================================================
def save_results(meta_results, trans_results, pairs):
    """Sauvegarde les résultats en JSON, TXT et génère les visualisations."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── JSON (données brutes) ──
    results = {
        "timestamp": timestamp,
        "n_letters": len(pairs),
        "metadata": meta_results,
        "transcription": trans_results
    }
    json_path = OUTPUT_DIR / f"evaluation_results_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n💾 JSON : {json_path}")

    # ── TXT (rapport lisible) ──
    txt_path = OUTPUT_DIR / f"evaluation_report_{timestamp}.txt"
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(f"=== Rapport d'évaluation Risala ===\n")
        f.write(f"Date        : {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"Lettres     : {len(pairs)}\n")
        f.write(f"Fichier     : {INPUT_FILE}\n\n")

        if meta_results:
            f.write("=== Fiabilité des métadonnées ===\n")
            for field, r in meta_results.items():
                f.write(f"  {field:20s}: {r['correct']:3d}/{r['total']:3d} ({r['pct']:.1f}%)\n")

        if trans_results:
            f.write("\n=== Transcription (CER / WER) ===\n")
            f.write(f"  CER global = {trans_results['cer_global']:.1f}%\n")
            f.write(f"  WER global = {trans_results['wer_global']:.1f}%\n")
            f.write(f"  Évalué sur {trans_results['n_evaluated']} lettre(s)\n")
            f.write("\n  Détail par lettre (CER) :\n")
            for lid, cer in sorted(trans_results["per_letter"].items(), key=lambda x: x[1], reverse=True):
                f.write(f"    {lid}: {cer:.1f}%\n")
    print(f"📄 TXT  : {txt_path}")

    # ── VISUALISATIONS ──
    if meta_results:
        _plot_metadata(meta_results, timestamp)
    if trans_results and trans_results["per_letter"]:
        _plot_cer_histogram(trans_results, timestamp)
        _plot_cer_vs_length(trans_results, timestamp)


def _plot_metadata(meta_results, timestamp):
    """Bar chart horizontal — fiabilité des métadonnées."""
    fields = list(meta_results.keys())
    pcts   = [meta_results[f]["pct"] for f in fields]
    colors = ["#2ecc71" if p >= 80 else "#e67e22" if p >= 50 else "#e74c3c" for p in pcts]

    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.barh(fields, pcts, color=colors, edgecolor="white", height=0.6)
    ax.set_xlim(0, 110)
    ax.set_xlabel("% corrects", fontsize=11)
    ax.set_title("Fiabilité des métadonnées (Manuel vs LLM)", fontsize=13, fontweight="bold")
    ax.axvline(x=80, color="gray", linestyle="--", linewidth=0.8, alpha=0.6)

    for bar, pct, f in zip(bars, pcts, fields):
        r = meta_results[f]
        ax.text(pct + 1, bar.get_y() + bar.get_height()/2,
                f"{pct:.1f}% ({r['correct']}/{r['total']})",
                va="center", fontsize=9)

    patches = [
        mpatches.Patch(color="#2ecc71", label="≥ 80%"),
        mpatches.Patch(color="#e67e22", label="50–79%"),
        mpatches.Patch(color="#e74c3c", label="< 50%"),
    ]
    ax.legend(handles=patches, loc="lower right", fontsize=9)
    plt.tight_layout()
    path = OUTPUT_DIR / f"fig1_metadata_{timestamp}.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"📊 Fig1 : {path}")


def _plot_cer_histogram(trans_results, timestamp):
    """Histogramme — distribution des CER par lettre."""
    cers = list(trans_results["per_letter"].values())
    bins = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]

    fig, ax = plt.subplots(figsize=(8, 5))
    n, _, patches = ax.hist(cers, bins=bins, color="#3498db", edgecolor="white", linewidth=0.8)
    for patch, left in zip(patches, bins):
        if left < 20:
            patch.set_facecolor("#2ecc71")
        elif left < 50:
            patch.set_facecolor("#e67e22")
        else:
            patch.set_facecolor("#e74c3c")

    ax.set_xlabel("CER (%)", fontsize=11)
    ax.set_ylabel("Nombre de lettres", fontsize=11)
    ax.set_title("Distribution du CER par lettre", fontsize=13, fontweight="bold")
    ax.axvline(x=trans_results["cer_global"], color="black", linestyle="--",
               linewidth=1.2, label=f"CER moyen = {trans_results['cer_global']:.1f}%")
    ax.legend(fontsize=9)
    ax.set_xticks(bins)
    plt.tight_layout()
    path = OUTPUT_DIR / f"fig2_cer_histogram_{timestamp}.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"📊 Fig2 : {path}")


def _plot_cer_vs_length(trans_results, timestamp):
    """Scatter plot — CER vs longueur de la transcription."""
    lids    = list(trans_results["per_letter"].keys())
    cers    = [trans_results["per_letter"][l] for l in lids]
    lengths = [trans_results["lengths"].get(l, 0) for l in lids]

    fig, ax = plt.subplots(figsize=(8, 5))
    sc = ax.scatter(lengths, cers, alpha=0.7, c=cers, cmap="RdYlGn_r",
                    vmin=0, vmax=100, edgecolors="white", linewidth=0.5, s=60)
    plt.colorbar(sc, ax=ax, label="CER (%)")
    ax.set_xlabel("Longueur de la transcription (caractères)", fontsize=11)
    ax.set_ylabel("CER (%)", fontsize=11)
    ax.set_title("CER vs longueur de la transcription", fontsize=13, fontweight="bold")

    # Annotations outliers
    for lid, cer, length in zip(lids, cers, lengths):
        if cer > 60:
            ax.annotate(lid, (length, cer), fontsize=7, alpha=0.8,
                        xytext=(4, 4), textcoords="offset points")
    plt.tight_layout()
    path = OUTPUT_DIR / f"fig3_cer_vs_length_{timestamp}.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"📊 Fig3 : {path}")


# ==============================================================================
# ÉVALUATION MÉTADONNÉES (version avec retour de données)
# ==============================================================================
def evaluate_metadata(pairs):
    print("\n=== Fiabilité des métadonnées ===")
    results = {}
    for field in METADATA_FIELDS:
        total, correct, skipped = 0, 0, 0
        for lid, ref_row, hyp_row in pairs:
            ref = ref_row.get(field)
            hyp = hyp_row.get(field)
            if ref:
                total += 1
                if normalize(ref) == normalize(hyp):
                    correct += 1
            elif hyp:
                skipped += 1
        if total == 0:
            print(f"  {field:20s}: aucune donnée dans Manuel")
            continue
        pct = correct / total * 100
        extra = f"  ({skipped} non relue(s))" if skipped else ""
        print(f"  {field:20s}: {correct:3d}/{total:3d} corrects ({pct:5.1f}%){extra}")
        results[field] = {"correct": correct, "total": total, "pct": pct, "skipped": skipped}
    return results


# ==============================================================================
# ÉVALUATION TRANSCRIPTION (version avec retour de données)
# ==============================================================================
def evaluate_transcription(pairs):
    print("\n=== Transcription (CER / WER) ===")
    ids, refs, hyps, lengths = [], [], [], []
    for lid, ref_row, hyp_row in pairs:
        ref = clean_text(ref_row.get("transcription"))
        hyp = clean_text(hyp_row.get("transcription"))
        if not ref:
            continue
        ids.append(lid)
        refs.append(ref)
        hyps.append(hyp)
        lengths.append(len(ref))

    if not refs:
        print("  Aucune transcription de référence dans Manuel.")
        return None

    cer_global = jiwer.cer(refs, hyps,
                           reference_transform=TEXT_TRANSFORM,
                           hypothesis_transform=TEXT_TRANSFORM) * 100
    wer_global = jiwer.wer(refs, hyps,
                           reference_transform=TEXT_TRANSFORM,
                           hypothesis_transform=TEXT_TRANSFORM) * 100

    print(f"  Évalué sur {len(refs)} lettre(s)")
    print(f"  CER global = {cer_global:.1f}%")
    print(f"  WER global = {wer_global:.1f}%")

    per_letter = {}
    per_lengths = {}
    print("\n  Détail par lettre (CER, du plus élevé au plus bas) :")
    for lid, r, h, ln in zip(ids, refs, hyps, lengths):
        c = jiwer.cer([r], [h],
                      reference_transform=TEXT_TRANSFORM,
                      hypothesis_transform=TEXT_TRANSFORM) * 100
        per_letter[lid] = c
        per_lengths[lid] = ln
        print(f"    {lid}: CER = {c:.1f}%")

    return {
        "cer_global": cer_global,
        "wer_global": wer_global,
        "n_evaluated": len(refs),
        "per_letter": per_letter,
        "lengths": per_lengths,
    }


def main():
    print(f"Lecture de {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

    if SHEET_MANUEL not in wb.sheetnames:
        print(f"⚠ Feuille '{SHEET_MANUEL}' introuvable.")
        sys.exit(1)
    if SHEET_LLM not in wb.sheetnames:
        print(f"⚠ Feuille '{SHEET_LLM}' introuvable.")
        sys.exit(1)

    manuel = load_manuel(wb)
    llm    = load_llm(wb)

    print(f"  Manuel : {len(manuel)} lettres (après regroupement)")
    print(f"  LLM    : {len(llm)} lettres")

    pairs = []
    for key, ref_row in manuel.items():
        if key in llm:
            pairs.append((f"{key[0]}_{key[1]}", ref_row, llm[key]))

    print(f"  Lettres en commun : {len(pairs)}")

    if not pairs:
        print("\n⚠ Aucune correspondance entre Manuel et LLM.")
        sys.exit(1)

    meta_results  = evaluate_metadata(pairs)  if MODE in ("metadata",  "all") else None
    trans_results = evaluate_transcription(pairs) if MODE in ("transcription", "all") else None

    save_results(meta_results or {}, trans_results, pairs)

if __name__ == "__main__":
    main()
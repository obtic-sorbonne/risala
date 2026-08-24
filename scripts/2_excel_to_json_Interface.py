"""
Convertit excel de drive (feuille Manuel) en lettres.json valide pour envoyer ds l'interface Risala.

Nouvelle structure : une ligne par page (L1_CONT_1, L1_ENV_1, etc.)
Le script regroupe les pages par lettre (L1, L2...) et :
  - Métadonnées : première valeur non vide trouvée dans toutes les pages (ENV en priorité)
  - Transcription : concaténation des valeurs non vides des pages CONT dans l'ordre


"""

import openpyxl
import json
import re
from datetime import datetime
from collections import defaultdict

# ── CONFIG ──────────────────────────────────────────────
INPUT_FILE  = "lettres.xlsx"
OUTPUT_FILE = "lettres_ar.json"
SHEET_NAME  = "Manuel"
# ────────────────────────────────────────────────────────

def fmt_date(val):
    """Convertit une valeur Excel en string YYYY-MM-DD ou None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        if val.year < 1940:
            return None
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s.lower() in ("nul", "none", "null", "", "#n/a", "nat"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year < 1940:
                return None
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s if s else None

def clean(val):
    """Nettoie une valeur : None si vide ou 'nul', str sinon."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("nul", "none", "null", "", "#n/a", "nan", "nat"):
        return None
    return s

def extract_letter_prefix(letter_id):
    """Extrait le préfixe de la lettre depuis letter_id.
    Ex: L1_CONT_1 → L1, L10_ENV_2 → L10, L1 → L1
    """
    if not letter_id:
        return None
    s = str(letter_id).strip()
    match = re.match(r'^(L\d+)', s, re.IGNORECASE)
    return match.group(1) if match else s

def is_cont(letter_id):
    """Retourne True si la page est une page de contenu (CONT)."""
    return bool(re.search(r'CONT', str(letter_id), re.IGNORECASE))

def detect_language(sender_name, transcription):
    """Détecte si la lettre est en arabe."""
    text = str(sender_name or "") + str(transcription or "")
    if re.search(r'[\u0600-\u06FF]', text):
        return "ar"
    return "fr"

def make_uid(folder_id, letter_prefix):
    """Construit un identifiant unique pour la lettre."""
    f = str(folder_id).strip() if folder_id else "UNKNOWN"
    return f"{f}_{letter_prefix}"

# ── MAIN ────────────────────────────────────────────────
print(f"Lecture de {INPUT_FILE}...")
wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

if SHEET_NAME not in wb.sheetnames:
    print(f"  ⚠ Feuille '{SHEET_NAME}' introuvable. Feuilles disponibles : {wb.sheetnames}")
    exit(1)

ws = wb[SHEET_NAME]
rows = list(ws.iter_rows(min_row=2, values_only=True))

# ── REGROUPEMENT DES PAGES PAR LETTRE ────────────────────
# Structure : { (folder_id, letter_prefix): [liste de rows] }
groups = defaultdict(list)

for row in rows:
    if not any(v is not None for v in row):
        continue
    try:
        (num, folder_id, letter_id, sender_name, sender_place,
         date_sent, recipient_name, recipient_place, date_received,
         inmate_id, transcription, notes) = row
    except ValueError:
        continue

    if not folder_id or not letter_id:
        continue

    prefix = extract_letter_prefix(letter_id)
    if not prefix:
        continue

    groups[(str(folder_id).strip(), prefix)].append({
        "letter_id":      str(letter_id).strip(),
        "sender_name":    clean(sender_name),
        "sender_place":   clean(sender_place),
        "date_sent":      fmt_date(date_sent),
        "recipient_name": clean(recipient_name),
        "recipient_place":clean(recipient_place),
        "date_received":  fmt_date(date_received),
        "inmate_id":      clean(inmate_id),
        "transcription":  clean(transcription),
        "notes":          clean(notes),
        "is_cont":        is_cont(letter_id),
    })

# ── CONSTRUCTION DES LETTRES JSON ────────────────────────
all_letters = []

for (folder_id, prefix), pages in groups.items():

    uid = make_uid(folder_id, prefix)

    # Métadonnées : première valeur non vide parmi toutes les pages
    def first(field):
        for p in pages:
            if p[field]:
                return p[field]
        return None

    # Transcription : concaténation des pages CONT dans l'ordre
    cont_pages = [p for p in pages if p["is_cont"] and p["transcription"]]
    transcription = " ".join(p["transcription"] for p in cont_pages)

    sender_name    = first("sender_name")
    transcription_clean = transcription if transcription.strip() else ""

    lang = detect_language(sender_name, transcription_clean)

    letter = {
        "id": uid,
        "type": "letter",
        "metadata": {
            "sender": {
                "id":   f"P_{uid}_S",
                "name": sender_name
            },
            "recipient": {
                "id":   f"P_{uid}_R",
                "name": first("recipient_name")
            },
            "date": {
                "sent":        first("date_sent"),
                "received":    first("date_received"),
                "approximate": False
            },
            "places": {
                "sender_place": {
                    "name":      first("sender_place"),
                    "latitude":  None,
                    "longitude": None
                },
                "recipient_place": {
                    "name":      first("recipient_place"),
                    "latitude":  None,
                    "longitude": None
                }
            },
            "correspondence_type": "sent",
            "language": lang,
            "source":   folder_id,
            "inmate_id": first("inmate_id"),
            "notes":    first("notes"),
            "entities": {
                "persons": [],
                "places":  [],
                "objects": []
            }
        },
        "content": {
            "text":   transcription_clean,
            "tokens": [],
            "lemmas": []
        },
        "analysis": {
            "keywords":   [],
            "topics":     [],
            "sentiments": []
        }
    }

    all_letters.append(letter)

# Tri par date puis par id
all_letters.sort(key=lambda l: (
    l["metadata"]["date"]["sent"] or "9999",
    l["id"]
))

output = {
    "metadata": {
        "title":   "Corpus de correspondances",
        "date":    datetime.today().strftime("%Y-%m-%d"),
        "creator": "bdd_to_json.py"
    },
    "letters": all_letters
}

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n✓ {len(all_letters)} lettres exportées → {OUTPUT_FILE}")
print("  Le fichier est du JSON valide, sans virgules traînantes, encodé UTF-8.")
